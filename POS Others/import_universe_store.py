import os
import sys
import csv
from typing import Dict, List, Tuple, Optional

import sqlite3
from db import get_connection

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# -----------------------------
# Helpers
# -----------------------------
def norm(s: str) -> str:
    """Normalize header / key strings (case-insensitive, no extra spaces)."""
    return " ".join((s or "").strip().lower().split())


def clean_val(v) -> str:
    """Turn cell value into a clean string."""
    if v is None:
        return ""
    return str(v).strip()


def read_rows_from_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for r in reader:
            rows.append({k: clean_val(v) for k, v in (r or {}).items()})
        return rows


def read_rows_from_xlsx(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed, cannot read .xlsx")

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # read header
    header = []
    for cell in ws[1]:
        header.append(clean_val(cell.value))

    rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, h in enumerate(header):
            if not h:
                continue
            d[h] = clean_val(row[i] if i < len(row) else "")
        # skip fully empty rows
        if any(v.strip() for v in d.values()):
            rows.append(d)
    return rows


def map_columns(row: Dict[str, str]) -> Dict[str, str]:
    """
    Maps user spreadsheet headers to canonical keys.
    Accepts minor header variations (case/spacing).
    """
    # Normalize all keys once
    kmap = {norm(k): k for k in row.keys()}

    def get_any(*candidates: str) -> str:
        for c in candidates:
            ck = norm(c)
            if ck in kmap:
                return row.get(kmap[ck], "").strip()
        return ""

    return {
        "order_booker_name": get_any("Order Booker Name", "OrderBooker Name", "OB Name"),
        "pjp_name": get_any("PJP Name", "PJP"),
        "store_name": get_any("Store Name", "Customer", "Customer Name", "Shop Name"),
        "owner_contact": get_any("Owner Contact #", "Owner Contact", "Contact", "Phone", "Phone #", "Mobile"),
        "address": get_any("Address", "Store Address", "Customer Address"),
    }


# -----------------------------
# DB upsert helpers
# -----------------------------
def get_or_create_order_booker(cur: sqlite3.Cursor, name: str) -> int:
    """
    Schema requires contact/address NOT NULL in order_bookers. :contentReference[oaicite:1]{index=1}
    We don’t have those columns in your import file, so we store empty strings.
    """
    cur.execute("SELECT id FROM order_bookers WHERE name = ? AND is_active = 1 LIMIT 1", (name,))
    r = cur.fetchone()
    if r:
        return r["id"]

    cur.execute(
        "INSERT INTO order_bookers (name, contact, address, is_active) VALUES (?, ?, ?, 1)",
        (name, "", ""),
    )
    return cur.lastrowid


def get_or_create_pjp(cur: sqlite3.Cursor, ob_id: int, pjp_name: str) -> int:
    """
    Schema requires day_of_week NOT NULL in pjps. :contentReference[oaicite:2]{index=2}
    Not present in import, so we set it to 'N/A'.
    """
    cur.execute(
        "SELECT id FROM pjps WHERE order_booker_id = ? AND pjp_name = ? AND is_active = 1 LIMIT 1",
        (ob_id, pjp_name),
    )
    r = cur.fetchone()
    if r:
        return r["id"]

    cur.execute(
        "INSERT INTO pjps (order_booker_id, pjp_name, day_of_week, is_active) VALUES (?, ?, ?, 1)",
        (ob_id, pjp_name, "N/A"),
    )
    return cur.lastrowid


def get_or_create_customer(cur: sqlite3.Cursor, pjp_id: int, store_name: str, contact: str, address: str) -> int:
    """
    Schema requires contact/address NOT NULL in customers. :contentReference[oaicite:3]{index=3}
    We’ll store empty string if missing.
    """
    contact = contact or ""
    address = address or ""

    # Try to avoid duplicates: same pjp + name (+ contact if present)
    if contact.strip():
        cur.execute(
            "SELECT id FROM customers WHERE pjp_id = ? AND name = ? AND contact = ? AND is_active = 1 LIMIT 1",
            (pjp_id, store_name, contact),
        )
    else:
        cur.execute(
            "SELECT id FROM customers WHERE pjp_id = ? AND name = ? AND is_active = 1 LIMIT 1",
            (pjp_id, store_name),
        )
    r = cur.fetchone()
    if r:
        return r["id"]

    cur.execute(
        "INSERT INTO customers (pjp_id, name, contact, address, is_active) VALUES (?, ?, ?, ?, 1)",
        (pjp_id, store_name, contact, address),
    )
    return cur.lastrowid


# -----------------------------
# Main import
# -----------------------------
def import_file(path: str, sheet_name: Optional[str] = None, dry_run: bool = False) -> None:
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        raw_rows = read_rows_from_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raw_rows = read_rows_from_xlsx(path, sheet_name=sheet_name)
    else:
        raise ValueError("Unsupported file type. Use .csv or .xlsx")

    conn = get_connection()
    cur = conn.cursor()

    created_ob = 0
    created_pjp = 0
    created_cust = 0
    skipped = 0

    # Quick caches to reduce DB hits
    ob_cache: Dict[str, int] = {}
    pjp_cache: Dict[Tuple[int, str], int] = {}

    for idx, rr in enumerate(raw_rows, start=1):
        mapped = map_columns(rr)

        ob_name = mapped["order_booker_name"]
        pjp_name = mapped["pjp_name"]
        store_name = mapped["store_name"]
        owner_contact = mapped["owner_contact"]
        address = mapped["address"]

        # Basic validation (must have these to insert meaningful rows)
        if not ob_name or not pjp_name or not store_name:
            skipped += 1
            continue

        # Order booker
        if ob_name in ob_cache:
            ob_id = ob_cache[ob_name]
        else:
            # detect if new before insert
            cur.execute("SELECT id FROM order_bookers WHERE name = ? AND is_active = 1 LIMIT 1", (ob_name,))
            existed = cur.fetchone() is not None
            ob_id = get_or_create_order_booker(cur, ob_name)
            ob_cache[ob_name] = ob_id
            if not existed:
                created_ob += 1

        # PJP
        pkey = (ob_id, pjp_name)
        if pkey in pjp_cache:
            pjp_id = pjp_cache[pkey]
        else:
            cur.execute(
                "SELECT id FROM pjps WHERE order_booker_id = ? AND pjp_name = ? AND is_active = 1 LIMIT 1",
                (ob_id, pjp_name),
            )
            existed = cur.fetchone() is not None
            pjp_id = get_or_create_pjp(cur, ob_id, pjp_name)
            pjp_cache[pkey] = pjp_id
            if not existed:
                created_pjp += 1

        # Customer
        # detect if new before insert (same logic used by get_or_create)
        if owner_contact.strip():
            cur.execute(
                "SELECT id FROM customers WHERE pjp_id = ? AND name = ? AND contact = ? AND is_active = 1 LIMIT 1",
                (pjp_id, store_name, owner_contact),
            )
        else:
            cur.execute(
                "SELECT id FROM customers WHERE pjp_id = ? AND name = ? AND is_active = 1 LIMIT 1",
                (pjp_id, store_name),
            )
        existed = cur.fetchone() is not None
        _ = get_or_create_customer(cur, pjp_id, store_name, owner_contact, address)
        if not existed:
            created_cust += 1

    if dry_run:
        conn.rollback()
        print("DRY RUN: no changes were saved.")
    else:
        conn.commit()

    conn.close()

    print("Import complete.")
    print(f"Created Order Bookers: {created_ob}")
    print(f"Created PJPs:         {created_pjp}")
    print(f"Created Customers:    {created_cust}")
    print(f"Skipped rows:         {skipped}")


if __name__ == "__main__":
    # Usage:
    #   python import_universe_store.py "Universe Store.csv"
    #   python import_universe_store.py "Universe Store.xlsx"
    #   python import_universe_store.py "Universe Store.xlsx" --sheet "Sheet1"
    #   python import_universe_store.py "Universe Store.xlsx" --dry-run
    args = sys.argv[1:]
    if not args:
        print("Usage: python import_universe_store.py <file.csv|file.xlsx> [--sheet SHEETNAME] [--dry-run]")
        sys.exit(1)

    path = args[0]
    sheet = None
    dry = False

    if "--sheet" in args:
        i = args.index("--sheet")
        if i + 1 < len(args):
            sheet = args[i + 1]

    if "--dry-run" in args:
        dry = True

    import_file(path, sheet_name=sheet, dry_run=dry)
